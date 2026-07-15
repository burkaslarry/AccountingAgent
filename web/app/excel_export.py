"""Excel export helpers."""

from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .agent import ReceiptRow
from .csv_export import CSV_HEADERS

DATE_FORMAT = "yyyy-mm-dd"
AMOUNT_FORMAT = "#,##0.00"


def rows_to_xlsx_bytes(rows: Iterable[ReceiptRow]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Receipts"

    header_font = Font(bold=True)
    for column_index, header in enumerate(CSV_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = header_font

    for row_index, row in enumerate(rows, start=2):
        sheet.cell(row=row_index, column=1, value=row.payment_date)
        sheet.cell(row=row_index, column=2, value=row.category)
        amount_cell = sheet.cell(row=row_index, column=3, value=float(row.amount))
        amount_cell.number_format = AMOUNT_FORMAT
        sheet.cell(row=row_index, column=4, value=row.currency)
        sheet.cell(row=row_index, column=5, value=row.reference_file)

    for column_index, header in enumerate(CSV_HEADERS, start=1):
        column_letter = get_column_letter(column_index)
        max_length = len(header)
        for cell in sheet[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 48)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
